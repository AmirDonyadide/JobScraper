"""
LinkedIn Job Scraper
====================
Uses Apify's LinkedIn Jobs Scraper actor to search for jobs
across multiple keywords with predefined filters, then deduplicates
and exports results locally, to Google Sheets, or both.

Requirements:
    pip install -r requirements.txt

Usage:
    1. Put APIFY_API_TOKEN in .env (or set it as an environment variable)
    2. Optional: set JOBSCRAPER_SOURCES to linkedin, indeed, or both
    3. Optional: set JOBSCRAPER_OUTPUT_MODE to excel, google_sheets, or both
    4. Run: python linkedin_job_scraper.py
    5. Open the printed output path or Google Sheet URL
"""

import json
import os
import time
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIGURATION — edit these
# ─────────────────────────────────────────────

TOKEN_ENV_VAR = "APIFY_API_TOKEN"
TOKEN_FILE = Path(__file__).with_name(".env")
TOKEN_PLACEHOLDER = "apify_api_XXXXXXXXXXXX"
GOOGLE_CLIENT_SECRET_FILE = Path(__file__).with_name("google_client_secret.json")
GOOGLE_TOKEN_FILE = Path(__file__).with_name("google_token.json")
GOOGLE_SPREADSHEET_ID_FILE = Path(__file__).with_name("google_spreadsheet_id.txt")
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
]


def load_local_env() -> dict[str, str]:
    """
    Load simple KEY=value settings from local .env.

    Supported .env formats:
        APIFY_API_TOKEN=apify_api_XXXXXXXXXXXX
        export APIFY_API_TOKEN=apify_api_XXXXXXXXXXXX
    """
    values = {}
    if not TOKEN_FILE.exists():
        return values

    for line in TOKEN_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export "):].strip()
        if "=" not in line:
            continue

        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip("\"'")

    return values


LOCAL_ENV = load_local_env()


def load_setting(name: str, default: str = "") -> str:
    return os.environ.get(name, LOCAL_ENV.get(name, default)).strip()


def load_int_setting(name: str, default: int) -> int:
    value = load_setting(name, str(default))
    try:
        return int(value)
    except ValueError:
        print(f"⚠ Invalid integer for {name}='{value}', using {default}.")
        return default


def load_bool_setting(name: str, default: bool) -> bool:
    value = load_setting(name, str(default).lower()).lower()
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    print(f"⚠ Invalid boolean for {name}='{value}', using {default}.")
    return default


def load_apify_token() -> str:
    """Load the Apify token from the environment first, then from local .env."""
    return load_setting(TOKEN_ENV_VAR)


APIFY_API_TOKEN = load_apify_token()
GOOGLE_SPREADSHEET_ID = load_setting("GOOGLE_SPREADSHEET_ID")
SCRAPER_TIMEZONE = load_setting("JOBSCRAPER_TIMEZONE", "Europe/Rome")
POSTED_TIMEZONE = load_setting("JOBSCRAPER_POSTED_TIMEZONE", "Europe/Berlin")

try:
    SCRAPER_TZ = ZoneInfo(SCRAPER_TIMEZONE)
except ZoneInfoNotFoundError as e:
    raise RuntimeError(
        f"Timezone '{SCRAPER_TIMEZONE}' is not available. "
        "Install tzdata or set JOBSCRAPER_TIMEZONE to a valid IANA timezone."
    ) from e

try:
    POSTED_TZ = ZoneInfo(POSTED_TIMEZONE)
except ZoneInfoNotFoundError as e:
    raise RuntimeError(
        f"Timezone '{POSTED_TIMEZONE}' is not available. "
        "Install tzdata or set JOBSCRAPER_POSTED_TIMEZONE to a valid IANA timezone."
    ) from e

RUN_STARTED_AT_UTC = datetime.now(timezone.utc)
RUN_STARTED_AT = RUN_STARTED_AT_UTC.astimezone(SCRAPER_TZ)
RUN_SHEET_NAME = RUN_STARTED_AT.strftime("%Y-%m-%d %H-%M-%S")

# Apify actors. Apify's raw REST API uses "~" between username and actor name.
LINKEDIN_ACTOR_ID = "curious_coder~linkedin-jobs-scraper"
INDEED_ACTOR_ID = "misceres~indeed-scraper"

# Max jobs to fetch per search URL (increase if you want more, costs more credits)
MAX_RESULTS_PER_SEARCH = load_int_setting("JOBSCRAPER_MAX_RESULTS_PER_SEARCH", 500)
INDEED_MAX_RESULTS_PER_SEARCH = load_int_setting("INDEED_MAX_RESULTS_PER_SEARCH", MAX_RESULTS_PER_SEARCH)

# Run multiple keyword searches at the same time. This is the main speed knob.
SEARCH_CONCURRENCY = max(1, load_int_setting("JOBSCRAPER_SEARCH_CONCURRENCY", 14))

# Apify settings for the child actor runs started by this script.
APIFY_RUN_MEMORY_MB = max(128, load_int_setting("APIFY_RUN_MEMORY_MB", 512))
APIFY_RUN_TIMEOUT_SECONDS = max(60, load_int_setting("APIFY_RUN_TIMEOUT_SECONDS", 300))
APIFY_CLIENT_TIMEOUT_SECONDS = max(
    APIFY_RUN_TIMEOUT_SECONDS + 30,
    load_int_setting("APIFY_CLIENT_TIMEOUT_SECONDS", APIFY_RUN_TIMEOUT_SECONDS + 60),
)

# Optional seconds to wait before starting another keyword request.
DELAY_BETWEEN_REQUESTS = max(0, load_int_setting("JOBSCRAPER_DELAY_BETWEEN_REQUESTS", 0))

# Choose: "linkedin", "indeed", or "both"
SOURCE_MODE = load_setting("JOBSCRAPER_SOURCES", "linkedin").lower()
SOURCE_DISPLAY_NAMES = {
    "linkedin": "LinkedIn",
    "indeed": "Indeed",
}

# Choose: "excel", "google_sheets", or "both"
OUTPUT_MODE = load_setting("JOBSCRAPER_OUTPUT_MODE", "excel").lower()

# Local Excel output path
EXCEL_OUTPUT_FILE = Path(__file__).with_name("jobs.xlsx")

# Output Google Sheet title
SPREADSHEET_TITLE = "jobs"

# ─────────────────────────────────────────────
#  SEARCH FILTERS (LinkedIn URL parameters)
# ─────────────────────────────────────────────
# f_TPR=r86400 -> Posted in last 24 hours
# f_E values   -> 1=Internship, 2=Entry level
# f_JT values  -> F=Full-time, P=Part-time, I=Internship

LOCATION = "Germany"
GEO_ID = "101282230"
PUBLISHED_AT = "r86400"
EXPERIENCE_LEVELS = ["1", "2"]
CONTRACT_TYPES = ["F", "P", "I"]
SCRAPE_COMPANY_DETAILS = load_bool_setting("JOBSCRAPER_SCRAPE_COMPANY_DETAILS", False)
USE_INCOGNITO_MODE = load_bool_setting("JOBSCRAPER_USE_INCOGNITO_MODE", True)
SPLIT_BY_LOCATION = load_bool_setting("JOBSCRAPER_SPLIT_BY_LOCATION", False)
SPLIT_COUNTRY = "DE"
EXCLUDED_TITLE_TERMS = ["Werkstudent"]

INDEED_COUNTRY = load_setting("INDEED_COUNTRY", "DE").upper()
INDEED_LOCATION = load_setting("INDEED_LOCATION", LOCATION)
INDEED_MAX_CONCURRENCY = load_int_setting("INDEED_MAX_CONCURRENCY", 5)
INDEED_SAVE_ONLY_UNIQUE_ITEMS = load_bool_setting("INDEED_SAVE_ONLY_UNIQUE_ITEMS", True)

# ─────────────────────────────────────────────
#  KEYWORDS
# ─────────────────────────────────────────────

KEYWORDS = [
    "3D Mapping",
    "Bauvermessung",
    "Cartography",
    "Earth Observation",
    "Erdbeobachtung",
    "Fernerkundung",
    "GeoAI",
    "Geoanalytics",
    "Geodaten",
    "Geodatenanalyse",
    "Geodatenbank",
    "Geodatenmanagement",
    "Geodäsie",
    "Geoinformatik",
    "Geoinformatiker",
    "Geoinformationssysteme",
    "Geomatik",
    "Geomatics",
    "Geomatiker",
    "Geospatial",
    "Geovisualisierung",
    "GIS",
    "GIS Fachkraft",
    "GIS Analyst",
    "GIS Entwickler",
    "GIS Spezialist",
    "Kartografie",
    "Laserscanning",
    "Photogrammetrie",
    "Raumdaten",
    "Remote Sensing",
    "Land Surveying",
    "Topografie",
    "Trassierung",
    "Vermessung",
    "Vermessungsingenieur",
    "Vermessungstechniker",
]

# ─────────────────────────────────────────────
#  APIFY API CALL
# ─────────────────────────────────────────────

class ApifyConfigurationError(RuntimeError):
    """Raised when Apify rejects the token, actor access, or paid actor setup."""


def apify_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {APIFY_API_TOKEN}"}


def apify_error_message(response: requests.Response) -> str:
    try:
        data = response.json()
    except ValueError:
        return response.text.strip()[:500] or response.reason

    if isinstance(data, dict):
        error = data.get("error")
        if isinstance(error, dict) and error.get("message"):
            return str(error["message"])
        if data.get("message"):
            return str(data["message"])

    return str(data)[:500]


INDEED_DOMAIN_BY_COUNTRY = {
    "us": "www.indeed.com",
    "gb": "uk.indeed.com",
    "uk": "uk.indeed.com",
    "de": "de.indeed.com",
    "at": "at.indeed.com",
    "ch": "ch.indeed.com",
    "fr": "fr.indeed.com",
    "nl": "nl.indeed.com",
    "it": "it.indeed.com",
    "es": "es.indeed.com",
    "ca": "ca.indeed.com",
    "au": "au.indeed.com",
}


def indeed_base_url() -> str:
    country_key = INDEED_COUNTRY.lower()
    domain = INDEED_DOMAIN_BY_COUNTRY.get(country_key, f"{country_key}.indeed.com")
    return f"https://{domain}"


def build_linkedin_search_url(keyword: str) -> str:
    params = {
        "keywords": keyword,
        "location": LOCATION,
        "geoId": GEO_ID,
        "f_TPR": PUBLISHED_AT,
        "f_E": ",".join(EXPERIENCE_LEVELS),
        "f_JT": ",".join(CONTRACT_TYPES),
        "position": "1",
        "pageNum": "0",
    }
    return f"https://www.linkedin.com/jobs/search/?{urlencode(params)}"


def build_linkedin_actor_input(search_url: str) -> dict:
    payload = {
        "urls": [search_url],
        "count": MAX_RESULTS_PER_SEARCH,
        "scrapeCompany": SCRAPE_COMPANY_DETAILS,
        "useIncognitoMode": USE_INCOGNITO_MODE,
        "splitByLocation": SPLIT_BY_LOCATION,
    }
    if SPLIT_BY_LOCATION:
        payload["splitCountry"] = SPLIT_COUNTRY
    return payload


def build_indeed_actor_input(keyword: str) -> dict:
    return {
        "country": INDEED_COUNTRY,
        "location": INDEED_LOCATION,
        "maxConcurrency": INDEED_MAX_CONCURRENCY,
        "maxItems": INDEED_MAX_RESULTS_PER_SEARCH,
        "position": keyword,
        "saveOnlyUniqueItems": INDEED_SAVE_ONLY_UNIQUE_ITEMS,
    }


def parse_job_sources() -> list[str]:
    aliases = {
        "linkedin": {"linkedin"},
        "li": {"linkedin"},
        "indeed": {"indeed"},
        "both": {"linkedin", "indeed"},
        "all": {"linkedin", "indeed"},
    }
    selected = aliases.get(SOURCE_MODE)
    if not selected:
        print(f"⚠ Unknown JOBSCRAPER_SOURCES '{SOURCE_MODE}', using LinkedIn only.")
        selected = {"linkedin"}

    return [source for source in ("linkedin", "indeed") if source in selected]


def get_searches(sources: list[str]) -> tuple[str, list[dict]]:
    searches = []

    if "linkedin" in sources:
        for keyword in KEYWORDS:
            searches.append({
                "source": "linkedin",
                "source_label": "LinkedIn",
                "keyword": keyword,
                "display_label": f"LinkedIn / {keyword}",
                "actor_id": LINKEDIN_ACTOR_ID,
                "payload": build_linkedin_actor_input(build_linkedin_search_url(keyword)),
                "max_items": MAX_RESULTS_PER_SEARCH,
            })

    if "indeed" in sources:
        for keyword in KEYWORDS:
            searches.append({
                "source": "indeed",
                "source_label": "Indeed",
                "keyword": keyword,
                "display_label": f"Indeed / {keyword}",
                "actor_id": INDEED_ACTOR_ID,
                "payload": build_indeed_actor_input(keyword),
                "max_items": INDEED_MAX_RESULTS_PER_SEARCH,
            })

    source_labels = ", ".join(SOURCE_DISPLAY_NAMES.get(source, source.title()) for source in sources)
    return f"generated {source_labels} keyword URLs", searches


def annotate_jobs(jobs: list[dict], source: str, source_label: str) -> list[dict]:
    annotated = []
    for job in jobs:
        job_copy = dict(job)
        job_copy["_source"] = source
        job_copy["_source_label"] = source_label
        annotated.append(job_copy)
    return annotated


def run_actor(actor_id: str, payload: dict, max_items: int) -> list[dict]:
    url = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset-items"
    params = {
        "timeout": APIFY_RUN_TIMEOUT_SECONDS,
        "memory": APIFY_RUN_MEMORY_MB,
        "maxItems": max_items,
    }

    response = requests.post(
        url,
        params=params,
        headers=apify_headers(),
        json=payload,
        timeout=APIFY_CLIENT_TIMEOUT_SECONDS,
    )

    if response.status_code in (401, 403):
        message = apify_error_message(response)
        raise ApifyConfigurationError(
            "Apify rejected the request. Check that APIFY_API_TOKEN is valid, "
            f"that your account can run {actor_id}, and that billing/trial "
            f"access is active. Apify said: {message}"
        )

    response.raise_for_status()
    data = response.json()

    # Apify returns a list directly for run-sync-get-dataset-items.
    if isinstance(data, list):
        return data
    # Some actors wrap results in a dict.
    if isinstance(data, dict) and "items" in data:
        return data["items"]
    return []


def fetch_jobs_for_search(search: dict) -> list[dict]:
    """
    Calls the configured Apify actor synchronously
    and returns annotated job dicts for the given search.
    """
    label = search["display_label"]
    try:
        jobs = run_actor(search["actor_id"], search["payload"], search["max_items"])
        return annotate_jobs(jobs, search["source"], search["source_label"])
    except requests.exceptions.Timeout:
        print(f"  ⚠ Timeout for search '{label}' — skipping.")
        return []
    except requests.exceptions.ConnectionError as e:
        raise ApifyConfigurationError(
            f"Could not connect to Apify API while searching '{label}'. "
            f"Check your internet connection/DNS and try again. Details: {e}"
        ) from e
    except ApifyConfigurationError:
        raise
    except requests.exceptions.HTTPError as e:
        response = e.response
        details = f" {apify_error_message(response)}" if response is not None else ""
        print(f"  ⚠ HTTP error for search '{label}': {e}.{details} — skipping.")
        return []
    except Exception as e:
        print(f"  ⚠ Unexpected error for search '{label}': {e} — skipping.")
        return []


def run_all_searches(searches: list[dict]) -> tuple[list[tuple[str, list]], list[str], dict[str, str], list[str]]:
    """
    Run keyword searches concurrently while preserving the original result order.

    Parallelizing here overlaps separate Apify actor runs. This mostly improves
    wall-clock time, while APIFY_RUN_MEMORY_MB controls each child run's memory.
    """
    all_results: list[tuple[int, str, list]] = []
    zero_searches: list[str] = []
    failed_sources: dict[str, str] = {}
    skipped_searches: list[str] = []

    max_workers = min(SEARCH_CONCURRENCY, len(searches))
    submitted_count = 0

    print(f"\nRunning up to {max_workers} search(es) in parallel ...")

    def submit_next(executor, search_iter, in_flight):
        nonlocal submitted_count
        while len(in_flight) < max_workers:
            try:
                idx, search = next(search_iter)
            except StopIteration:
                return

            label = search["display_label"]
            source = search["source"]
            source_label = search["source_label"]

            if source in failed_sources:
                print(f"\n[{idx:02d}/{len(searches)}] Search: '{label}'")
                print(f"  → Skipped because {source_label} failed earlier in this run.")
                skipped_searches.append(label)
                continue

            if DELAY_BETWEEN_REQUESTS and submitted_count:
                time.sleep(DELAY_BETWEEN_REQUESTS)

            print(f"\n[{idx:02d}/{len(searches)}] Search: '{label}'")
            print(f"  → Calling Apify actor {search['actor_id']} ...")
            future = executor.submit(fetch_jobs_for_search, search)
            in_flight[future] = (idx, search)
            submitted_count += 1

    search_iter = iter(enumerate(searches, start=1))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        in_flight = {}
        submit_next(executor, search_iter, in_flight)

        while in_flight:
            for future in as_completed(tuple(in_flight)):
                idx, search = in_flight.pop(future)
                label = search["display_label"]
                keyword = search["keyword"]
                source = search["source"]
                source_label = search["source_label"]

                try:
                    jobs = future.result()
                except ApifyConfigurationError as e:
                    if source not in failed_sources:
                        print(f"\n❌ {e}")
                        print(
                            f"   Continuing with any results from other sources. "
                            f"Remaining {source_label} searches will be skipped."
                        )
                        failed_sources[source] = str(e)
                    skipped_searches.append(label)
                else:
                    all_results.append((idx, keyword, jobs))
                    if jobs:
                        print(f"  ✓ Completed '{label}': {len(jobs)} job(s) found")
                    else:
                        print(f"  — Completed '{label}': 0 results")
                        zero_searches.append(label)

                submit_next(executor, search_iter, in_flight)
                break

    ordered_results = [
        (keyword, jobs)
        for _, keyword, jobs in sorted(all_results, key=lambda item: item[0])
    ]
    return ordered_results, zero_searches, failed_sources, skipped_searches


# ─────────────────────────────────────────────
#  DEDUPLICATION
# ─────────────────────────────────────────────

def make_dedup_key(job: dict) -> str:
    """
    Primary key: source + jobId (if available).
    Fallback key: title + company + location (lowercased, stripped).
    """
    source = str(job.get("_source") or "unknown").lower().strip()
    job_id = job.get("jobId") or job.get("job_id") or job.get("id") or ""
    if job_id:
        return f"{source}|{str(job_id).strip()}"

    title = get_title(job).lower().strip()
    company = get_company(job).lower().strip()
    location = get_location(job).lower().strip()
    return f"{source}|{title}|{company}|{location}"


def merge_and_deduplicate(all_results: list[tuple[str, list]]) -> list[dict]:
    """
    all_results: [(keyword, [job, job, ...]), ...]
    Returns a deduplicated list of jobs, each annotated with
    'keywords_matched' = list of keywords that found it.
    """
    seen: dict[str, dict] = {}   # dedup_key → job dict

    for keyword, jobs in all_results:
        for job in jobs:
            key = make_dedup_key(job)
            if key in seen:
                # Job already recorded — just add keyword to its matched list
                seen[key]["keywords_matched"].append(keyword)
            else:
                job_copy = dict(job)
                job_copy["keywords_matched"] = [keyword]
                seen[key] = job_copy

    return list(seen.values())


# ─────────────────────────────────────────────
#  FIELD EXTRACTION HELPERS
# ─────────────────────────────────────────────

def safe(job: dict, *keys) -> str:
    """Try multiple key names and return first non-empty value."""
    for k in keys:
        v = job.get(k)
        if v and str(v).strip():
            return str(v).strip()
    return "N/A"


def nested(job: dict, *keys) -> str:
    value = job
    for key in keys:
        if not isinstance(value, dict):
            return "N/A"
        value = value.get(key)
    if value is None or value == "":
        return "N/A"
    return sheet_safe(value)


def get_source_label(job: dict) -> str:
    return safe(job, "_source_label")


def get_title(job: dict) -> str:
    return safe(job, "title", "positionName", "jobTitle", "job_title", "name")


def get_company(job: dict) -> str:
    return field(job, "companyName", "company", "organization", "jobSourceName") if "companyDetails" not in job else first_value(
        nested(job, "companyDetails", "name"),
        field(job, "companyName", "company", "organization", "jobSourceName"),
    )


def get_location(job: dict) -> str:
    if isinstance(job.get("location"), dict):
        return first_value(
            nested(job, "location", "formatted", "long"),
            nested(job, "location", "fullAddress"),
            nested(job, "location", "city"),
        )
    return field(job, "location", "formattedLocation", "jobLocation", "place")


def first_value(*values: str) -> str:
    for value in values:
        if value and value != "N/A":
            return value
    return "N/A"


def get_job_url(job: dict) -> str:
    url = (
        job.get("jobUrl")
        or job.get("job_url")
        or job.get("linkedinUrl")
        or job.get("linkedin_url")
        or job.get("url")
        or job.get("link")
        or ""
    )
    if url:
        return url

    view_job_link = job.get("viewJobLink") or ""
    if view_job_link:
        if str(view_job_link).startswith("http"):
            return str(view_job_link)
        return f"{indeed_base_url()}{view_job_link}"

    job_id = job.get("jobId") or job.get("job_id") or job.get("id") or ""
    if job_id:
        if job.get("_source") == "indeed":
            return f"{indeed_base_url()}/viewjob?jk={job_id}"
        return f"https://www.linkedin.com/jobs/view/{job_id}/"
    return "N/A"


def get_posted(job: dict) -> str:
    posted_keys = (
        "postedAt", "posted_at", "publishedAt", "published_at",
        "datePosted", "date_posted", "posted", "listedAt", "listed_at",
    )
    fallback = ""
    for key in posted_keys:
        value = job.get(key)
        if value and str(value).strip():
            posted_at = parse_datetime_value(value)
            if posted_at:
                return format_posted_datetime(posted_at)
            if not fallback:
                fallback = sheet_safe(value)

    pub_date = job.get("pubDate")
    posted_at = parse_datetime_value(pub_date)
    if posted_at:
        return format_posted_datetime(posted_at)
    if fallback:
        return fallback
    return format_posted_value(pub_date)


def get_job_type(job: dict) -> str:
    return field(job, "employmentType", "employment_type", "jobType", "job_type", "contractType", "contract_type", "type", "jobTypes")


def get_job_type_choice(job: dict) -> str:
    job_type = get_job_type(job).casefold()
    if "part" in job_type:
        return "part-time"
    if "full" in job_type:
        return "full-time"
    return ""


def get_experience(job: dict) -> str:
    return safe(job, "experienceLevel", "experience_level", "seniorityLevel", "seniority_level", "seniority")


def get_experience_choice(job: dict) -> str:
    experience = get_experience(job).casefold()
    if "entry" in experience:
        return "entry level"
    return "not applicable"


def get_apply_url(job: dict) -> str:
    return field(job, "applyUrl", "apply_url", "originalApplyUrl", "thirdPartyApplyUrl", "externalApplyLink")


def get_company_website(job: dict) -> str:
    return first_value(
        field(job, "companyWebsite", "company_website", "website"),
        nested(job, "companyDetails", "websiteUrl"),
    )


def format_timestamp(value) -> str:
    posted_at = parse_datetime_value(value)
    if posted_at:
        return format_posted_datetime(posted_at)
    return "N/A"


def parse_datetime_value(value):
    if value in (None, ""):
        return None
    try:
        timestamp = float(value)
    except (TypeError, ValueError):
        timestamp = None

    if timestamp is not None:
        if timestamp > 10_000_000_000:
            timestamp = timestamp / 1000
        return datetime.fromtimestamp(timestamp, timezone.utc).astimezone(POSTED_TZ)

    text = str(value).strip()
    if not text or text == "N/A":
        return None

    iso_text = text
    if iso_text.endswith("Z"):
        iso_text = iso_text[:-1] + "+00:00"

    try:
        parsed = datetime.fromisoformat(iso_text)
    except ValueError:
        return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=POSTED_TZ)
    return parsed.astimezone(POSTED_TZ)


def format_posted_datetime(posted_at: datetime) -> str:
    return posted_at.strftime("%Y-%m-%d %H:%M:%S")


def format_posted_value(value) -> str:
    posted_at = parse_datetime_value(value)
    if posted_at:
        return format_posted_datetime(posted_at)
    return sheet_safe(value)


def has_excluded_title(job: dict) -> bool:
    title = get_title(job).casefold()
    return any(term.casefold() in title for term in EXCLUDED_TITLE_TERMS)


def filter_excluded_titles(jobs: list[dict]) -> tuple[list[dict], int]:
    filtered_jobs = [job for job in jobs if not has_excluded_title(job)]
    return filtered_jobs, len(jobs) - len(filtered_jobs)


# ─────────────────────────────────────────────
#  GOOGLE SHEETS EXPORT
# ─────────────────────────────────────────────

HEADER = [
    "Application Status", "App", "Job Title", "Company", "Location",
    "Job Type", "Posted", "Experience Level", "Applicants",
    "Keywords Matched", "Job URL", "Apply URL",
]

APPLICATION_STATUS_OPTIONS = ["applied", "rejected", "interview", "accepted"]
JOB_TYPE_OPTIONS = ["part-time", "full-time"]
EXPERIENCE_LEVEL_OPTIONS = ["entry level", "not applicable"]

MAX_CELL_CHARS = 49000


class GoogleSheetsExportError(RuntimeError):
    """Raised when Google Sheets export is not configured correctly."""


def sheet_safe(value) -> str:
    if value is None or value == "":
        return "N/A"
    if isinstance(value, list):
        text = ", ".join(str(item) for item in value if item is not None and str(item).strip())
    elif isinstance(value, dict):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)

    text = text.strip()
    if not text:
        return "N/A"
    if len(text) > MAX_CELL_CHARS:
        return text[:MAX_CELL_CHARS - 20] + " ... [truncated]"
    if text[0] in "=+-@":
        return "'" + text
    return text


def field(job: dict, *keys) -> str:
    for key in keys:
        value = job.get(key)
        if value is not None and sheet_safe(value) != "N/A":
            return sheet_safe(value)
    return "N/A"


def sheets_string(value: str) -> str:
    return str(value).replace('"', '""')


def hyperlink_formula(url: str, label: str) -> str:
    if not url or url == "N/A":
        return "N/A"
    return f'=HYPERLINK("{sheets_string(url)}", "{sheets_string(label)}")'


def make_job_rows(jobs: list[dict]) -> list[list]:
    rows = [HEADER]
    for i, job in enumerate(jobs, start=1):
        job_url = get_job_url(job)
        apply_url = get_apply_url(job)
        rows.append([
            "",
            get_source_label(job),
            get_title(job),
            get_company(job),
            get_location(job),
            get_job_type_choice(job),
            get_posted(job),
            get_experience_choice(job),
            field(job, "applicantsCount", "applicants_count"),
            ", ".join(job.get("keywords_matched", [])),
            hyperlink_formula(job_url, "Open Job"),
            hyperlink_formula(apply_url, "Open Apply"),
        ])
    return rows


def unique_name(existing_names: set[str], base_name: str, max_length: int | None = None) -> str:
    name = base_name[:max_length] if max_length else base_name
    if name not in existing_names:
        return name

    counter = 2
    while True:
        suffix = f" ({counter})"
        if max_length:
            candidate = f"{base_name[:max_length - len(suffix)]}{suffix}"
        else:
            candidate = f"{base_name}{suffix}"
        if candidate not in existing_names:
            return candidate
        counter += 1


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────

COLOR_HEADER_BG = "102C53"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ODD = "CADCFC"
COLOR_ROW_EVEN = "FFFFFF"
COLOR_ACCENT = "C9A84C"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_cell(cell):
    cell.font = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_data_cell(cell, row_idx: int, is_url: bool = False):
    bg = COLOR_ROW_ODD if row_idx % 2 == 1 else COLOR_ROW_EVEN
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER
    if is_url:
        cell.font = Font(color=COLOR_ACCENT, name="Calibri", size=10, underline="single")
    else:
        cell.font = Font(name="Calibri", size=10)


def parse_hyperlink_formula(value: str):
    prefix = '=HYPERLINK("'
    separator = '", "'
    suffix = '")'
    if not isinstance(value, str) or not value.startswith(prefix) or not value.endswith(suffix):
        return None

    body = value[len(prefix):-len(suffix)]
    if separator not in body:
        return None

    url, label = body.split(separator, 1)
    return url.replace('""', '"'), label.replace('""', '"')


def excel_value(value):
    parsed = parse_hyperlink_formula(value)
    if parsed:
        return parsed[1]
    return value


def export_to_excel(jobs: list[dict], filename: Path) -> str:
    if filename.exists():
        wb = openpyxl.load_workbook(filename)
        sheet_name = unique_name(set(wb.sheetnames), RUN_SHEET_NAME, 31)
        ws = wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = unique_name(set(), RUN_SHEET_NAME, 31)

    job_rows = make_job_rows(jobs)
    for row in job_rows:
        ws.append([excel_value(value) for value in row])

    for cell in ws[1]:
        style_header_cell(cell)
    ws.row_dimensions[1].height = 30

    url_columns = {idx for idx, name in enumerate(HEADER, start=1) if "URL" in name}
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            style_data_cell(cell, row_idx, is_url=col_idx in url_columns)
            parsed = parse_hyperlink_formula(job_rows[row_idx - 1][col_idx - 1])
            if parsed:
                cell.hyperlink = parsed[0]

    width_by_header = {
        "Application Status": 20,
        "App": 14,
        "Job Title": 40,
        "Company": 32,
        "Location": 28,
        "Job Type": 18,
        "Posted": 22,
        "Experience Level": 22,
        "Keywords Matched": 32,
        "Job URL": 18,
        "Apply URL": 18,
    }
    for col_idx, header in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_by_header.get(header, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    return f"{filename} (sheet: {ws.title})"


# ─────────────────────────────────────────────
#  GOOGLE SHEETS EXPORT
# ─────────────────────────────────────────────

def build_google_sheets_service():
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as e:
        raise GoogleSheetsExportError(
            "Missing Google API packages. Install them with:\n"
            "   pip install -r requirements.txt"
        ) from e

    creds = None
    if GOOGLE_TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(GOOGLE_TOKEN_FILE), GOOGLE_SCOPES)
        if not creds.has_scopes(GOOGLE_SCOPES):
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not GOOGLE_CLIENT_SECRET_FILE.exists():
                raise GoogleSheetsExportError(
                    f"Missing {GOOGLE_CLIENT_SECRET_FILE.name}. Create a Google OAuth "
                    "Desktop client, download its JSON credentials, and save it in this folder."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(GOOGLE_CLIENT_SECRET_FILE), GOOGLE_SCOPES)
            creds = flow.run_local_server(port=0)

        GOOGLE_TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("sheets", "v4", credentials=creds)


def quote_sheet_name(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def update_values(service, spreadsheet_id: str, sheet_name: str, rows: list[list]):
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows},
    ).execute()


def header_index(name: str) -> int:
    return HEADER.index(name)


def column_range(sheet_id: int, column_name: str, end_row_index: int) -> dict:
    column_idx = header_index(column_name)
    return {
        "sheetId": sheet_id,
        "startRowIndex": 1,
        "endRowIndex": end_row_index,
        "startColumnIndex": column_idx,
        "endColumnIndex": column_idx + 1,
    }


def dropdown_validation_request(sheet_id: int, column_name: str, options: list[str], end_row_index: int) -> dict:
    return {
        "setDataValidation": {
            "range": column_range(sheet_id, column_name, end_row_index),
            "rule": {
                "condition": {
                    "type": "ONE_OF_LIST",
                    "values": [{"userEnteredValue": option} for option in options],
                },
                "showCustomUi": True,
                "strict": True,
            },
        }
    }


def date_time_format_request(sheet_id: int, column_name: str, end_row_index: int) -> dict:
    return {
        "repeatCell": {
            "range": column_range(sheet_id, column_name, end_row_index),
            "cell": {
                "userEnteredFormat": {
                    "numberFormat": {
                        "type": "DATE_TIME",
                        "pattern": "yyyy-mm-dd hh:mm",
                    }
                }
            },
            "fields": "userEnteredFormat.numberFormat",
        }
    }


def format_spreadsheet(service, spreadsheet_id: str, sheet_id: int, job_row_count: int):
    editable_row_count = max(job_row_count, 1000)
    requests_body = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {"frozenRowCount": 1},
                },
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.063, "green": 0.173, "blue": 0.325},
                        "horizontalAlignment": "CENTER",
                        "textFormat": {
                            "bold": True,
                            "foregroundColor": {"red": 1, "green": 1, "blue": 1},
                        },
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,textFormat)",
            }
        },
        {
            "setBasicFilter": {
                "filter": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": max(job_row_count, 1),
                        "startColumnIndex": 0,
                        "endColumnIndex": len(HEADER),
                    }
                }
            }
        },
        {
            "autoResizeDimensions": {
                "dimensions": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": len(HEADER),
                }
            }
        },
        dropdown_validation_request(
            sheet_id,
            "Application Status",
            APPLICATION_STATUS_OPTIONS,
            editable_row_count,
        ),
        dropdown_validation_request(
            sheet_id,
            "Job Type",
            JOB_TYPE_OPTIONS,
            editable_row_count,
        ),
        dropdown_validation_request(
            sheet_id,
            "Experience Level",
            EXPERIENCE_LEVEL_OPTIONS,
            editable_row_count,
        ),
        date_time_format_request(sheet_id, "Posted", editable_row_count),
    ]

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": requests_body},
    ).execute()


def read_google_spreadsheet_id() -> str:
    if GOOGLE_SPREADSHEET_ID:
        return GOOGLE_SPREADSHEET_ID
    if GOOGLE_SPREADSHEET_ID_FILE.exists():
        return GOOGLE_SPREADSHEET_ID_FILE.read_text(encoding="utf-8").strip()
    return ""


def get_google_spreadsheet(service, spreadsheet_id: str) -> dict:
    return service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="spreadsheetId,spreadsheetUrl,sheets(properties(sheetId,title))",
    ).execute()


def create_google_spreadsheet(service) -> tuple[dict, str, int]:
    spreadsheet = service.spreadsheets().create(
        body={
            "properties": {"title": SPREADSHEET_TITLE},
            "sheets": [{"properties": {"title": RUN_SHEET_NAME}}],
        },
        fields="spreadsheetId,spreadsheetUrl,sheets(properties(sheetId,title))",
    ).execute()
    GOOGLE_SPREADSHEET_ID_FILE.write_text(spreadsheet["spreadsheetId"], encoding="utf-8")
    sheet_id = spreadsheet["sheets"][0]["properties"]["sheetId"]
    return spreadsheet, RUN_SHEET_NAME, sheet_id


def add_google_run_sheet(service, spreadsheet_id: str, existing_names: set[str]) -> tuple[str, int]:
    sheet_name = unique_name(existing_names, RUN_SHEET_NAME)
    response = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "addSheet": {
                        "properties": {"title": sheet_name},
                    }
                }
            ]
        },
    ).execute()
    sheet_id = response["replies"][0]["addSheet"]["properties"]["sheetId"]
    return sheet_name, sheet_id


def get_or_create_google_run_sheet(service) -> tuple[str, str, str, int]:
    spreadsheet_id = read_google_spreadsheet_id()
    if not spreadsheet_id:
        spreadsheet, sheet_name, sheet_id = create_google_spreadsheet(service)
        return spreadsheet["spreadsheetId"], spreadsheet["spreadsheetUrl"], sheet_name, sheet_id

    try:
        spreadsheet = get_google_spreadsheet(service, spreadsheet_id)
    except Exception as e:
        raise GoogleSheetsExportError(
            f"Could not open Google spreadsheet ID '{spreadsheet_id}'. "
            f"Check {GOOGLE_SPREADSHEET_ID_FILE.name}, or delete it to create a new 'jobs' spreadsheet. "
            f"Details: {e}"
        ) from e

    existing_names = {
        sheet["properties"]["title"]
        for sheet in spreadsheet.get("sheets", [])
    }
    sheet_name, sheet_id = add_google_run_sheet(service, spreadsheet_id, existing_names)
    return spreadsheet_id, spreadsheet["spreadsheetUrl"], sheet_name, sheet_id


def export_to_google_sheets(service, jobs: list[dict]) -> str:
    spreadsheet_id, spreadsheet_url, sheet_name, sheet_id = get_or_create_google_run_sheet(service)
    job_rows = make_job_rows(jobs)

    update_values(service, spreadsheet_id, sheet_name, job_rows)
    format_spreadsheet(service, spreadsheet_id, sheet_id, len(job_rows))

    return f"{spreadsheet_url} (sheet: {sheet_name})"


def parse_output_mode() -> set[str]:
    aliases = {
        "excel": {"excel"},
        "local": {"excel"},
        "xlsx": {"excel"},
        "google": {"google_sheets"},
        "drive": {"google_sheets"},
        "google_sheets": {"google_sheets"},
        "sheets": {"google_sheets"},
        "both": {"excel", "google_sheets"},
        "all": {"excel", "google_sheets"},
    }
    modes = aliases.get(OUTPUT_MODE)
    if modes:
        return modes

    print(f"⚠ Unknown OUTPUT_MODE '{OUTPUT_MODE}', using local Excel output.")
    return {"excel"}


def format_duration(seconds: float) -> str:
    total_seconds = int(round(seconds))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    if hours:
        return f"{hours}h {minutes}m {seconds}s"
    if minutes:
        return f"{minutes}m {seconds}s"
    return f"{seconds}s"


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    run_started = time.perf_counter()

    if not APIFY_API_TOKEN or APIFY_API_TOKEN == TOKEN_PLACEHOLDER:
        print(f"❌ Please set {TOKEN_ENV_VAR} in {TOKEN_FILE.name} or as an environment variable.")
        print(f"   Add this line to {TOKEN_FILE.name}:")
        print(f"   {TOKEN_ENV_VAR}={TOKEN_PLACEHOLDER}")
        return

    job_sources = parse_job_sources()
    search_source, searches = get_searches(job_sources)
    if not searches:
        print("❌ No searches configured.")
        print("   Add keywords in the script.")
        return

    output_modes = parse_output_mode()
    google_sheets_service = None
    if "google_sheets" in output_modes:
        print("Checking Google Sheets access ...")
        try:
            google_sheets_service = build_google_sheets_service()
            print("Google Sheets access OK.")
        except GoogleSheetsExportError as e:
            print(f"\n❌ {e}")
            return

    print("=" * 60)
    print(f"  Job Scraper — {RUN_STARTED_AT.strftime('%Y-%m-%d %H:%M %Z')}")
    print(f"  UTC start time: {RUN_STARTED_AT_UTC.strftime('%Y-%m-%d %H:%M UTC')}")
    print(f"  Sources: {', '.join(SOURCE_DISPLAY_NAMES.get(source, source.title()) for source in job_sources)}")
    if "linkedin" in job_sources:
        print(f"  LinkedIn actor: {LINKEDIN_ACTOR_ID}")
    if "indeed" in job_sources:
        print(f"  Indeed actor: {INDEED_ACTOR_ID}")
    print(f"  Search source: {search_source}")
    print(f"  Output mode: {', '.join(sorted(output_modes))}")
    print(f"  Timezone: {SCRAPER_TIMEZONE}")
    print(f"  Posted timezone: {POSTED_TIMEZONE}")
    print(f"  Searches: {len(searches)}")
    print(f"  Search concurrency: {SEARCH_CONCURRENCY}")
    print(f"  Apify child run memory: {APIFY_RUN_MEMORY_MB} MB")
    print(f"  Apify child run timeout: {APIFY_RUN_TIMEOUT_SECONDS}s")
    if DELAY_BETWEEN_REQUESTS:
        print(f"  Delay between starting searches: {DELAY_BETWEEN_REQUESTS}s")
    if "linkedin" in job_sources:
        print(f"  LinkedIn job types: {', '.join(CONTRACT_TYPES)}")
        print(f"  LinkedIn experience levels: {', '.join(EXPERIENCE_LEVELS)}")
        print(f"  LinkedIn max results/search: {MAX_RESULTS_PER_SEARCH}")
        print(f"  LinkedIn scrape company details: {SCRAPE_COMPANY_DETAILS}")
    if "indeed" in job_sources:
        print(f"  Indeed country/location: {INDEED_COUNTRY.upper()} / {INDEED_LOCATION}")
        print(f"  Indeed max results/search: {INDEED_MAX_RESULTS_PER_SEARCH}")
        print(f"  Indeed max concurrency: {INDEED_MAX_CONCURRENCY}")
        print(f"  Indeed save unique only: {INDEED_SAVE_ONLY_UNIQUE_ITEMS}")
    print("=" * 60)

    all_results, zero_searches, failed_sources, skipped_searches = run_all_searches(searches)

    # ── Deduplicate ──────────────────────────────
    print("\n" + "─" * 60)
    print("Deduplicating results ...")
    unique_jobs = merge_and_deduplicate(all_results)
    print(f"  → {len(unique_jobs)} unique job(s) after deduplication")

    # ── Final title exclusions ───────────────────
    print("Applying title filters ...")
    unique_jobs, excluded_title_count = filter_excluded_titles(unique_jobs)
    terms = ", ".join(EXCLUDED_TITLE_TERMS)
    print(f"  → Removed {excluded_title_count} job(s) containing: {terms}")

    # ── Sort by posted date (most recent first) ──
    # Keep N/A entries at the bottom
    print("Sorting results ...")
    def sort_key(job):
        posted = get_posted(job)
        return posted if posted != "N/A" else "0000"

    unique_jobs.sort(key=sort_key, reverse=True)

    # ── Export ───────────────────────────────────
    outputs = []
    if "excel" in output_modes:
        print(f"Saving local Excel file '{EXCEL_OUTPUT_FILE.name}' ...")
        outputs.append(("Excel file", export_to_excel(unique_jobs, EXCEL_OUTPUT_FILE)))

    if "google_sheets" in output_modes:
        print("Creating Google Sheet ...")
        try:
            spreadsheet_url = export_to_google_sheets(
                google_sheets_service,
                unique_jobs,
            )
        except GoogleSheetsExportError as e:
            print(f"\n❌ {e}")
            return
        outputs.append(("Google Sheet", spreadsheet_url))

    # ── Summary ──────────────────────────────────
    print("\n" + "=" * 60)
    print(f"  Searched {len(searches)} search URL(s) → Found {len(unique_jobs)} unique job postings.")
    print(f"  Total runtime: {format_duration(time.perf_counter() - run_started)}")
    if excluded_title_count:
        print(f"  Excluded by title rule: {excluded_title_count}")
    if zero_searches:
        print(f"  Searches with 0 results ({len(zero_searches)}):")
        for label in zero_searches:
            print(f"    • {label}")
    if failed_sources:
        print(f"  Failed source(s) ({len(failed_sources)}):")
        for source, message in failed_sources.items():
            source_label = SOURCE_DISPLAY_NAMES.get(source, source.title())
            print(f"    • {source_label}: {message[:180]}")
    if skipped_searches:
        print(f"  Skipped searches after source failure: {len(skipped_searches)}")
    for label, destination in outputs:
        print(f"  {label}: {destination}")
    print("=" * 60)


if __name__ == "__main__":
    main()
