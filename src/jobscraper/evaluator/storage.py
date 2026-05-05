"""Excel and Google Sheets storage adapters for evaluator results."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from jobscraper.evaluator.models import (
    DETAIL_COLUMNS,
    OUTPUT_COLUMNS,
    REMOVED_AI_OUTPUT_COLUMNS,
    EvaluationError,
    GoogleSheetsError,
    JobEvaluation,
)
from jobscraper.evaluator.parsing import normalize_header, trim_trailing_blank_headers
from jobscraper.google_sheets import build_google_sheets_service, quote_sheet_name
from jobscraper.paths import GOOGLE_SPREADSHEET_ID_FILE


def resolve_sheet_name(existing_names: list[str], requested: str) -> str:
    """Resolve a requested sheet name, with ``latest`` selecting the newest tab."""
    if not existing_names:
        raise EvaluationError("The workbook/spreadsheet has no sheets.")
    if not requested or requested == "latest":
        return existing_names[-1]
    if requested in existing_names:
        return requested
    raise EvaluationError(
        f"Sheet '{requested}' was not found. Available sheets: "
        f"{', '.join(existing_names)}"
    )


def read_excel_input(
    path: Path,
    requested_sheet: str,
) -> tuple[Any, Any, str, list[str], list[list[Any]]]:
    """Read headers and rows from an Excel workbook sheet."""
    if not path.exists():
        raise EvaluationError(f"Excel file not found: {path}")

    workbook = openpyxl.load_workbook(path)
    sheet_name = resolve_sheet_name(workbook.sheetnames, requested_sheet)
    worksheet = workbook[sheet_name]
    headers = trim_trailing_blank_headers(
        [
            worksheet.cell(row=1, column=col_idx).value
            for col_idx in range(1, worksheet.max_column + 1)
        ]
    )
    if not headers:
        raise EvaluationError(f"Sheet '{sheet_name}' has no header row.")

    rows = [
        [
            worksheet.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, len(headers) + 1)
        ]
        for row_idx in range(2, worksheet.max_row + 1)
    ]
    return workbook, worksheet, sheet_name, headers, rows


def columns_to_remove_after_evaluation(headers: list[str]) -> list[int]:
    """Return zero-based column indexes to delete after evaluator output is saved."""
    removable = {
        normalize_header(column)
        for column in [*REMOVED_AI_OUTPUT_COLUMNS, *DETAIL_COLUMNS]
    }
    return [
        idx
        for idx, header in enumerate(headers)
        if normalize_header(header) in removable
    ]


def remove_excel_columns_after_evaluation(worksheet: Any, headers: list[str]) -> None:
    """Delete legacy AI metadata and job detail columns from an Excel worksheet."""
    for column_idx in sorted(columns_to_remove_after_evaluation(headers), reverse=True):
        worksheet.delete_cols(column_idx + 1)


def write_excel_output(
    workbook: Any,
    worksheet: Any,
    path: Path,
    headers: list[str],
    header_map: dict[str, int],
    evaluations: dict[int, JobEvaluation],
) -> None:
    """Write evaluator columns and results back to an Excel worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx).value = header

    for evaluation in evaluations.values():
        for column in OUTPUT_COLUMNS:
            column_idx = header_map[normalize_header(column)] + 1
            worksheet.cell(
                row=evaluation.row_number,
                column=column_idx,
            ).value = evaluation.value_for_column(column)

    remove_excel_columns_after_evaluation(worksheet, headers)
    workbook.save(path)


def build_evaluator_google_sheets_service() -> Any:
    """Build a Google Sheets service for evaluator reads and writes."""
    return build_google_sheets_service(error_cls=GoogleSheetsError)


def read_google_spreadsheet_id(cli_value: str) -> str:
    """Resolve a spreadsheet ID from CLI, env, or local cache file."""
    if cli_value:
        return cli_value

    from jobscraper.env import EnvSettings

    env_value = EnvSettings().get("GOOGLE_SPREADSHEET_ID")
    if env_value:
        return env_value
    if GOOGLE_SPREADSHEET_ID_FILE.exists():
        return GOOGLE_SPREADSHEET_ID_FILE.read_text(encoding="utf-8").strip()
    return ""


def read_google_input(
    service: Any,
    spreadsheet_id: str,
    requested_sheet: str,
) -> tuple[str, list[str], list[list[Any]]]:
    """Read headers and rows from a Google Sheet tab."""
    metadata = (
        service.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, fields="sheets(properties(title))")
        .execute()
    )
    sheet_names = [sheet["properties"]["title"] for sheet in metadata.get("sheets", [])]
    sheet_name = resolve_sheet_name(sheet_names, requested_sheet)
    response = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=quote_sheet_name(sheet_name))
        .execute()
    )
    values = response.get("values", [])
    if not values:
        raise EvaluationError(f"Google Sheet tab '{sheet_name}' is empty.")

    headers = trim_trailing_blank_headers(values[0])
    if not headers:
        raise EvaluationError(f"Google Sheet tab '{sheet_name}' has no header row.")
    rows = [list(row) for row in values[1:]]
    return sheet_name, headers, rows


def write_google_output(
    service: Any,
    spreadsheet_id: str,
    sheet_name: str,
    headers: list[str],
    header_map: dict[str, int],
    evaluations: dict[int, JobEvaluation],
) -> None:
    """Write evaluator columns and results back to a Google Sheet tab."""
    data = []
    for column in OUTPUT_COLUMNS:
        column_idx = header_map[normalize_header(column)]
        column_letter = get_column_letter(column_idx + 1)
        data.append(
            {
                "range": f"{quote_sheet_name(sheet_name)}!{column_letter}1",
                "values": [[headers[column_idx]]],
            }
        )
        for evaluation in evaluations.values():
            data.append(
                {
                    "range": (
                        f"{quote_sheet_name(sheet_name)}!"
                        f"{column_letter}{evaluation.row_number}"
                    ),
                    "values": [[evaluation.value_for_column(column)]],
                }
            )

    for idx in range(0, len(data), 500):
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "valueInputOption": "RAW",
                "data": data[idx : idx + 500],
            },
        ).execute()

    remove_google_columns_after_evaluation(
        service,
        spreadsheet_id,
        sheet_name,
        headers,
    )


def get_google_sheet_id(service: Any, spreadsheet_id: str, sheet_name: str) -> int:
    """Return the numeric Google Sheets tab ID for a sheet title."""
    metadata = (
        service.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, fields="sheets(properties(sheetId,title))")
        .execute()
    )
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if properties.get("title") == sheet_name:
            return int(properties["sheetId"])
    raise EvaluationError(f"Google Sheet tab '{sheet_name}' was not found.")


def remove_google_columns_after_evaluation(
    service: Any,
    spreadsheet_id: str,
    sheet_name: str,
    headers: list[str],
) -> None:
    """Delete legacy AI metadata and job detail columns from a Google Sheet tab."""
    column_indexes = columns_to_remove_after_evaluation(headers)
    if not column_indexes:
        return

    sheet_id = get_google_sheet_id(service, spreadsheet_id, sheet_name)
    requests = [
        {
            "deleteDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": column_idx,
                    "endIndex": column_idx + 1,
                }
            }
        }
        for column_idx in sorted(column_indexes, reverse=True)
    ]
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": requests},
    ).execute()
