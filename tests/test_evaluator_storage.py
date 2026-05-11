"""Tests for evaluator result storage."""

from __future__ import annotations

import openpyxl

from jobfinder.evaluator.models import JobEvaluation
from jobfinder.evaluator.parsing import ensure_output_columns
from jobfinder.evaluator.storage import write_excel_output


def test_write_excel_output_can_skip_cleanup_for_incremental_save(tmp_path):
    """Incremental saves should preserve source columns until final cleanup."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    output_file = tmp_path / "jobs.xlsx"
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])

    worksheet.cell(row=2, column=1).value = "GIS Analyst"
    worksheet.cell(row=2, column=2).value = "Analyze spatial data"

    write_excel_output(
        workbook,
        worksheet,
        output_file,
        headers,
        header_map,
        {
            2: JobEvaluation(
                row_number=2,
                verdict="Suitable",
                fit_score=90,
                reason="Strong match.",
                tailored_cv="CV",
                model="test-model",
            )
        },
        cleanup_columns=False,
    )

    saved = openpyxl.load_workbook(output_file)
    saved_worksheet = saved.active
    assert saved_worksheet.cell(row=1, column=2).value == "Job Description"
    assert saved_worksheet.cell(row=2, column=3).value == "Suitable"

    write_excel_output(
        saved,
        saved_worksheet,
        output_file,
        headers,
        header_map,
        {},
        cleanup_columns=True,
    )

    finalized = openpyxl.load_workbook(output_file)
    finalized_headers = [
        finalized.active.cell(row=1, column=idx).value
        for idx in range(1, finalized.active.max_column + 1)
    ]
    assert "Job Description" not in finalized_headers
    assert finalized.active.cell(row=2, column=2).value == "Suitable"
